import time

from pytesseract import pytesseract
import openpyxl as xl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import os


wb = xl.load_workbook("upwork8.xlsx", data_only=True)
sheet = wb.worksheets[0]
ws = wb.worksheets[0]

sheet.cell(1, 1).value = "Car Name"
sheet.cell(1, 2).value = "Doors"
sheet.cell(1, 3).value = "Years"
sheet.cell(1, 4).value = "Engine"
sheet.cell(1, 5).value = "Battery"

filename = 'img2text.jpeg'
program_files_location = os.environ["ProgramFiles"]
path_to_tesseract = program_files_location + "/Tesseract-OCR/tesseract.exe"
pytesseract.tesseract_cmd = path_to_tesseract

# Passing the image object to image_to_string() function
# This function will extract the text from the image
text = str(pytesseract.image_to_string(filename)).replace("‘","")
lines = str(text).split("\n")
lines = list(filter(None,lines))
lines.pop(0)
lines.pop()

excel_row = 2

print(f'{"Car Name":^60} : {"Car Doors":^25} : {"Years":^15} : {"Engine":^20} : {"Battery":^20}')
print("=" * 147)

for line in lines:
    car_name = line.split(" ")
    for i, text in enumerate(car_name):
        if text == "Door":
            car_name = car_name[:i-1]
            car_name_hold = ""
            for j in car_name:
                car_name_hold += j + " "
            break
    car_name = car_name_hold
    car_door = line.split(" ")
    for i, text in enumerate(car_door):
        if text == "Door":
            car_door = str(car_door[i-1] + " " + text + " " + car_door[i+1]).replace("(","").replace(")","")
            break
    volts = ""
    if line.split(" ")[-3:][2].lower().__contains__("w"):
        volts = line.split(" ")[-3:][0] + " " + line.split(" ")[-3:][1] + " " + line.split(" ")[-3:][2]
    else:
        volts = line.split(" ")[-3:][1] + " " + line.split(" ")[-3:][2]

    liters = line.split(" ")
    liter = ""
    for i, word in enumerate(liters):
        if word.lower().__contains__("litre"):
            liter = liters[i-1] + " " + word + " " + liters[i+1]
            break
    years = ""
    split_lines = line.replace("~","").replace(">","").replace("  ", " ")
    for i, text in enumerate(split_lines.split(" ")):
        if text.lower().__contains__("door"):
            years = split_lines.split(" ")[i + 2] + " - " + split_lines.split(" ")[i + 3]
            break

    car_name_cell = sheet.cell(excel_row, 1)
    car_door_cell = sheet.cell(excel_row, 2)
    car_years_cell = sheet.cell(excel_row, 3)
    car_engine_cell = sheet.cell(excel_row, 4)
    car_battery_cell = sheet.cell(excel_row, 5)
    car_name_cell.value, car_door_cell.value, car_years_cell.value, car_engine_cell.value, car_battery_cell.value = car_name, car_door, years, liter, volts
    excel_row += 1
    print(f'{car_name:60} : {car_door:25} : {years:15} : {liter:20} : {volts:20}')

dim_holder = DimensionHolder(worksheet=ws)
dim_holder[get_column_letter(1)] = ColumnDimension(ws, min=1, max=1, width=70)
dim_holder[get_column_letter(2)] = ColumnDimension(ws, min=2, max=2, width=30)
dim_holder[get_column_letter(3)] = ColumnDimension(ws, min=3, max=3, width=20)
dim_holder[get_column_letter(4)] = ColumnDimension(ws, min=4, max=4, width=25)
dim_holder[get_column_letter(5)] = ColumnDimension(ws, min=4, max=4, width=40)
ws.column_dimensions = dim_holder

wb.save("upwork8.xlsx")
time.sleep(5)

os.startfile("upwork8.xlsx")


