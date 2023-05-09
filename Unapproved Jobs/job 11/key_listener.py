import pyautogui
import openpyxl

# Open the Excel workbook
wb = openpyxl.load_workbook('C:\\Users\\seant\\Downloads\\Contact_List.xlsx')
sheet = wb.active

# Track the last used row in the sheet
last_row = sheet.max_row

# Start listening for Ctrl+C key presses
while True:
    if pyautogui.hotkey('ctrl', 'c'):
        # Get the selected text from the clipboard
        text = pyautogui.hotkey('ctrl', 'v')

        # Insert the text into a new row in the Excel sheet
        sheet.cell(row=last_row+1, column=3).value = text
        wb.save('C:\\Users\\seant\\Downloads\\Contact_List.xlsx')

        # Move to the next row
        last_row += 1