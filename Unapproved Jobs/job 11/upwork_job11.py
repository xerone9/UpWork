import time
import keyboard as k
import openpyxl
import pyperclip

wb = openpyxl.load_workbook('C:\\Users\\HYSTOU\\Desktop\\Contact_List.xlsx')
sheet = wb.active

# Track the last used row in the sheet
last_row = sheet.max_row

# Start listening for Ctrl+C key presses
while True:
    if k.is_pressed("ctrl") and k.is_pressed("c"):
        time.sleep(0.5)
        text = pyperclip.paste()

        # Insert the text into a new row in the Excel sheet
        sheet.cell(row=last_row + 1, column=3).value = text
        wb.save('C:\\Users\\HYSTOU\\Desktop\\Contact_List.xlsx')

        # Move to the next row
        last_row += 1
        print(text)
