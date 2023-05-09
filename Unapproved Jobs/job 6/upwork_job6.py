import openpyxl as xl
import os

desktop = os.path.expanduser("~/Desktop/")
wb = xl.load_workbook(desktop + "Pages from Kamus kecil Bahasa Oirata.xlsx", data_only=True)
# wb = xl.load_workbook(r'C:\Users\HYSTOU\Desktop\sest.xlsx')
sheet = wb.worksheets[0]



for row in range(1, sheet.max_row + 1):
    excel = sheet.cell(row, 2).value
    name_cell = sheet.cell(row, 1)
    pronounciation_cell = sheet.cell(row, 2)
    parts_of_speech_cell = sheet.cell(row, 3)
    meaning_cell = sheet.cell(row, 4)
    if str(excel).__contains__("  /"):
        name = str(excel).split("  /")[0]
        pronounciation = name.split("/")[0]
        parts_of_speech = str(excel).split("/ ")[1].split(".")[0]
        meaning_combined = str(excel).split(".")[1:]
        meaning = ""
        for i in meaning_combined:
            meaning += i
        name_cell.value = name
        pronounciation_cell.value = pronounciation
        parts_of_speech_cell.value = parts_of_speech
        meaning_cell.value = meaning
    else:
        name_cell.value = excel
wb.save("Pages from Kamus kecil Bahasa Oirata.xlsx")