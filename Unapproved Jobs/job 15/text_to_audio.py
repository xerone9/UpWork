import openpyxl as xl
from gtts import gTTS

wb = xl.load_workbook('spanish\Spanish_words.xlsx', data_only=True)
sheet = wb.worksheets[0]

for row in range(1, sheet.max_row + 1):
    cell = sheet.cell(row, 1)
    word = str(cell.value)
    output_file = "spanish\\" + word + ".mp3"
    tts = gTTS(text=word, lang="es")
    tts.save(output_file)
    print(f"{word:10} :   Pronunciation saved as :   {output_file}")










