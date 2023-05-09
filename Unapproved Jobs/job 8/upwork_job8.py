from PyPDF2 import PdfFileReader
import openpyxl as xl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import os

wb = xl.load_workbook("upwork7.xlsx", data_only=True)
sheet = wb.worksheets[0]
ws = wb.worksheets[0]

# Headers
sheet.cell(1, 1).value = "Amount"
sheet.cell(1, 2).value = "Description"
sheet.cell(1, 3).value = "Dated"
sheet.cell(1, 4).value = "Receipt No"


pdf_file = "upwork.pdf"
reader = PdfFileReader(pdf_file)
pages = reader.numPages
extracted_text = ""
for i in range(pages):
    page = reader.getPage(i)
    extracted_text += str(page.extract_text())

receipt_no = extracted_text.split("Customer Copy No. ")[1].split("No. ")[0]
dated = extracted_text.split("DIEGOREPRINT\n")[1].split(" ")[0]
lines = extracted_text.split("\n")
print("")
print(f'{"Amount":^8} : {"Description":^120} : {"Dated":^12} : {"Receipt No":^15}')
print("=" * 160)

excel_row = 2
for i, line in enumerate(lines):
    if line.startswith("R") and line.__contains__("$"):
        description = line.split("EA")[1].split(" /")[0]
        amount = "$" + line.split("$")[-1]
        print(f'{amount:8} : {description:120} : {dated:12} : {receipt_no:15}')
        amount_cell = sheet.cell(excel_row, 1)
        description_cell = sheet.cell(excel_row, 2)
        dated_cell = sheet.cell(excel_row, 3)
        receipt_no_cell = sheet.cell(excel_row, 4)
        amount_cell.value, description_cell.value, dated_cell.value, receipt_no_cell.value = amount, description, dated, receipt_no
        excel_row += 1
    elif line.startswith("R"):
        if line.startswith("REF") or line.startswith("Reviewer"):
            pass
        else:
            line = str(line + lines[i+1])
            description = line.split("EA")[1].split(" /")[0]
            amount = "$" + line.split("$")[-1]
            print(f'{amount:8} : {description:120} : {dated:12} : {receipt_no:15}')
            amount_cell = sheet.cell(excel_row, 1)
            description_cell = sheet.cell(excel_row, 2)
            dated_cell = sheet.cell(excel_row, 3)
            receipt_no_cell = sheet.cell(excel_row, 4)
            amount_cell.value, description_cell.value, dated_cell.value, receipt_no_cell.value = amount, description, dated, receipt_no
            excel_row += 1


dim_holder = DimensionHolder(worksheet=ws)
dim_holder[get_column_letter(1)] = ColumnDimension(ws, min=1, max=1, width=10)
dim_holder[get_column_letter(2)] = ColumnDimension(ws, min=2, max=2, width=70)
dim_holder[get_column_letter(3)] = ColumnDimension(ws, min=3, max=3, width=15)
dim_holder[get_column_letter(4)] = ColumnDimension(ws, min=4, max=4, width=15)
ws.column_dimensions = dim_holder

wb.save("upwork7.xlsx")
os.startfile("upwork7.xlsx")



