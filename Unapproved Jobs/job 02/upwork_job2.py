import openpyxl as xl
import selenium.common.exceptions
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
# from selenium.common.exceptions import NoSuchElementException
# from selenium.common.exceptions import StaleElementReferenceException
# from selenium.webdriver.support import expected_conditions
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC

#


cities = ["Accomack County, Virginia", "Addison village, Illinois", "Adelanto city, California"]

PATH = "C:\Program Files (x86)\chromedriver.exe"
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

driver = webdriver.Chrome(options=chrome_options)

driver.get("https://www.google.com/search?q=srk&sxsrf=AJOqlzXHNrLu9sNmhi8ylYCDADBY3LNDUw%3A1677309115508&source=hp&ei=u7T5Y726HO6o9u8PiaqewAc&iflsig=AK50M_UAAAAAY_nCy56VGhat3rjUJJCtkEPYUjkjgFjd&ved=0ahUKEwj9g4awj7D9AhVulP0HHQmVB3gQ4dUDCAg&uact=5&oq=srk&gs_lcp=Cgdnd3Mtd2l6EAMyBAgjECcyBAgjECcyBAgjECcyCwguELEDENQCEJECMgUIABCRAjIECAAQQzILCC4QgAQQsQMQgwEyCwguEIAEELEDEIMBMggILhCABBCxAzIFCAAQgAQ6BwgjEOoCECc6CwgAEIAEELEDEIMBOggIABCABBCxA1CgA1icCGDYCWgBcAB4AIAB2gGIAaYDkgEDMi0ymAEAoAEBsAEK&sclient=gws-wiz")

wb = xl.load_workbook("Places 30-40k.xlsx", data_only=True)
sheet = wb.worksheets[0]
for row in range(68, sheet.max_row + 1):
    city_name = sheet.cell(row, 1).value
    print(city_name)
    search_google = driver.find_element(by=By.NAME, value="q")
    search_google.clear()
    search_google.send_keys(city_name)
    search_google.send_keys(Keys.ENTER)
    time.sleep(2)
    get_website = driver.find_element(by=By.TAG_NAME, value="cite")
    paste_website = sheet.cell(row, 4)
    paste_website.value = get_website.text
    wb.save("Places 30-40k.xlsx")
    print(get_website.text)


