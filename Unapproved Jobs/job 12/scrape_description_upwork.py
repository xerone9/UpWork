from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import selenium.common.exceptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import time
import openpyxl as xl

PATH = "C:\Program Files (x86)\chromedriver.exe"
chrome_options = Options()
chrome_options.add_argument("--disable-popup-blocking")
chrome_options.add_experimental_option("detach", True)

wb = xl.load_workbook("politesocietyshop_scrapped.xlsx", data_only=True)
sheet = wb.worksheets[0]

driver = webdriver.Chrome(options=chrome_options)
# URL = "https://politesocietyshop.com/products/boxy-t-shirt"

file_path = 'upwork_links.txt'
row = 74

# Read the file line by line
with open(file_path, 'r') as f:
    for line in f:
        try:
            URL = line.strip()

            driver.get(URL)

            summary_element = driver.find_element(By.CLASS_NAME, "cc-accordion-item__title")

            # Click on the summary element to expand the content
            summary_element.click()

            # Find the <p> element with the specified data-mce-fragment attribute
            p_element = driver.find_element(By.CSS_SELECTOR, "p[data-mce-fragment='1']")

            # Retrieve the text content using JavaScript
            p_text = driver.execute_script("return arguments[0].textContent;", p_element)

            # Print the text content of the <p> element
            last_row = sheet.max_row
            sheet.cell(row=row, column=4, value=p_text)
            wb.save("politesocietyshop_scrapped.xlsx")
            row += 1
        except selenium.common.exceptions.NoSuchElementException:
            row += 1
