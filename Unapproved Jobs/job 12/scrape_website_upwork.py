from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import selenium.common.exceptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import time
import os
import openpyxl as xl

wb = xl.load_workbook("politesocietyshop_scrapped.xlsx", data_only=True)
sheet = wb.worksheets[0]

PATH = "C:\Program Files (x86)\chromedriver.exe"
chrome_options = Options()
chrome_options.add_argument("--disable-popup-blocking")
chrome_options.add_experimental_option("detach", True)

driver = webdriver.Chrome(options=chrome_options)
URL = "https://politesocietyshop.com/collections/all"
driver.get(URL)


try:
    # Find the pagination element
    pagination = driver.find_element(By.CSS_SELECTOR, 'nav[aria-label="Pagination"] ul.pagination')

    # Find all page links within the pagination element
    page_links = pagination.find_elements(By.CSS_SELECTOR, 'li.pagination__page a.pagination__page-link')

    # Find the last page number
    last_page_number = int(page_links[-1].text)
    print(str(last_page_number))

    # Loop through pages
    for page_number in range(1, last_page_number + 1):
        # Visit each page
        page_url = f'{URL}?page={page_number}'
        driver.get(page_url)

        scroll_height = 1000
        num_scrolls = 10
        for _ in range(num_scrolls):
            driver.execute_script("window.scrollTo(0, {})".format(scroll_height))
            time.sleep(1)  # Allow time for elements to load

            scroll_height += 1000  # Adjust scroll height as needed

        # Wait for the parent elements to be present
        parent_elements = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, '.card.card--with-hover.column.third, .card.column.third'))
        )

        # Wait for the parent elements to be present
        parent_elements = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, '.card.card--with-hover.column.third, .card.column.third'))
        )

        for parent_element in parent_elements:
            # Find the card__image element within the parent element
            card_image = parent_element.find_element(By.CLASS_NAME, 'card__image')

            # Find the picture tag within the card__image element
            picture_tag = card_image.find_element(By.TAG_NAME, 'picture')

            # Find the source tag within the picture tag
            source_tag = picture_tag.find_element(By.TAG_NAME, 'source')

            # Get the image source
            image_source = source_tag.get_attribute('srcset').split()[0]

            # Find the card__content element within the parent element
            card_content = parent_element.find_element(By.CLASS_NAME, 'card__content')

            # Find the product title link
            product_title_link = card_content.find_element(By.CLASS_NAME, 'card__link')

            # Get the product URL
            product_url = product_title_link.get_attribute('href')

            # Get the product name (text content of the <a> element)
            product_name = product_title_link.text

            # Find the product price element
            get_to_product_price_element = card_content.find_element(By.CLASS_NAME, 'product-price')

            # Find the product price element
            product_price_element = get_to_product_price_element.find_element(By.CLASS_NAME, 'money')

            # Get the product price from the 'data-currency' attribute
            product_price = product_price_element.get_attribute('data-currency-usd')

            # Print the extracted data
            print("Image source:", image_source)
            print("Product URL:", product_url)
            print("Product Price:", product_price)
            print("Product Price:", product_name)
            print(" ")



            last_row = sheet.max_row
            excel_image_source = 3
            excel_product_URL = 6
            excel_product_price = 2
            excel_product_name = 1

            sheet.cell(row=last_row + 1, column=excel_image_source, value=image_source)
            sheet.cell(row=last_row + 1, column=excel_product_URL, value=product_url)
            sheet.cell(row=last_row + 1, column=excel_product_price, value=product_price)
            sheet.cell(row=last_row + 1, column=excel_product_name, value=product_name)

            wb.save("politesocietyshop_scrapped.xlsx")


except IndexError:
    pass